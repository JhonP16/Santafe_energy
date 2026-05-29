"""Servicio de catálogos: recursos de generación y agentes.

Lee los archivos Excel provistos en el reto (``recursos.xlsx`` y
``agentes.xlsx``) y expone:

* el mapeo de código SIC -> nombre legible (para enriquecer los datos de XM),
* el filtro de plantas de Despacho Central (DC) para el panel de generación,
* el listado de agentes comercializadores para el panel de demanda.

Los catálogos se cargan una sola vez y se mantienen en memoria.
"""
from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import openpyxl

from app.core.config import get_settings

# Fila donde comienzan los encabezados/datos en los Excel del reto.
_HEADER_ROW = 4
_FIRST_DATA_ROW = 5


def _fold(text: str | None) -> str:
    """Normaliza texto para comparaciones: sin acentos, en mayúsculas y sin espacios."""
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    without_accents = "".join(c for c in normalized if not unicodedata.combining(c))
    return without_accents.strip().upper()


@dataclass(frozen=True)
class Resource:
    """Recurso de generación del catálogo."""

    code: str
    name: str
    dispatch_type: str  # "DC" (Despacho Central) o "ND" (No Despachado)

    @property
    def is_centrally_dispatched(self) -> bool:
        return _fold(self.dispatch_type) == "DC"


@dataclass(frozen=True)
class Agent:
    """Agente del mercado del catálogo."""

    code: str
    name: str
    activity: str

    @property
    def is_marketer(self) -> bool:
        """Indica si el agente es comercializador."""
        return _fold(self.activity).startswith("COMERCIALIZ")


class Catalog:
    """Catálogos de recursos y agentes cargados desde los Excel."""

    def __init__(self, resources: dict[str, Resource], agents: dict[str, Agent]) -> None:
        self._resources = resources
        self._agents = agents

    # --- Recursos --------------------------------------------------------

    def resource_name(self, code: str) -> str:
        """Nombre del recurso; si no está en el catálogo, devuelve el propio código."""
        resource = self._resources.get(code)
        return resource.name if resource else code

    def is_centrally_dispatched(self, code: str) -> bool:
        """True si el recurso es una planta de Despacho Central (DC)."""
        resource = self._resources.get(code)
        return bool(resource and resource.is_centrally_dispatched)

    # --- Agentes ---------------------------------------------------------

    def agent_name(self, code: str) -> str:
        """Nombre del agente; si no está en el catálogo, devuelve el propio código."""
        agent = self._agents.get(code)
        return agent.name if agent else code

    def marketers(self) -> list[Agent]:
        """Agentes comercializadores, ordenados alfabéticamente por nombre."""
        marketers = [a for a in self._agents.values() if a.is_marketer]
        return sorted(marketers, key=lambda a: a.name)


def _load_resources(path: Path) -> dict[str, Resource]:
    """Carga el catálogo de recursos desde el Excel."""
    worksheet = _open_sheet(path)
    resources: dict[str, Resource] = {}
    for row in worksheet.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True):
        code = _clean(row[0])
        if not code:
            continue
        resources[code] = Resource(
            code=code,
            name=_clean(row[1]) or code,
            dispatch_type=_clean(row[5]),
        )
    return resources


def _load_agents(path: Path) -> dict[str, Agent]:
    """Carga el catálogo de agentes desde el Excel."""
    worksheet = _open_sheet(path)
    agents: dict[str, Agent] = {}
    for row in worksheet.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True):
        code = _clean(row[0])
        if not code:
            continue
        agents[code] = Agent(
            code=code,
            name=_clean(row[1]) or code,
            activity=_clean(row[2]),
        )
    return agents


def _open_sheet(path: Path):
    """Abre la hoja activa del Excel en modo solo lectura."""
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de catálogo: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return workbook.active


def _clean(value: object) -> str:
    """Convierte una celda a string limpio (sin espacios sobrantes)."""
    if value is None:
        return ""
    return str(value).strip()


@lru_cache
def get_catalog() -> Catalog:
    """Devuelve el catálogo cargado (una sola vez) desde los Excel configurados."""
    settings = get_settings()
    return Catalog(
        resources=_load_resources(settings.recursos_xlsx),
        agents=_load_agents(settings.agentes_xlsx),
    )
